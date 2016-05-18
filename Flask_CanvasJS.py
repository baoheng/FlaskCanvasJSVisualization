import json

from flask import Flask, render_template
from openpyxl import load_workbook

app = Flask(__name__)


@app.route('/')
@app.route('/index')
def index():
    info = chart_info()
    info_keys = list(info.keys())
    return render_template("index2.html",
                           title="Pie Chart for " + str(info_keys[len(info_keys)-1]),
                           chart_info=chart_info())


def input_file_by_year():
    f = load_workbook(filename="G:\PycharmProjects\hello_flask\static\Debt Affordability Study Data.xlsx", read_only=True, use_iterators=True)
    data_by_year = {}
    for row in f["Bond Issuance (Granular Data)"].iter_rows(row_offset=2):
        if row[3].value not in data_by_year.keys():
            data_by_year[row[3].value] = {row[6].value: row[5].value}
        else:
            if row[6].value not in data_by_year[row[3].value].keys():
                data_by_year[row[3].value][row[6].value] = row[5].value
            else:
                data_by_year[row[3].value][row[6].value] += row[5].value
    return data_by_year


def chart_info():
    data_by_year = input_file_by_year()
    data = {}

    for year in data_by_year.keys():
        dataDict = {"type": "pie",
                "showInLegend": True,
                "indexLabel": "{y}"}
        data_point = []
        for key in data_by_year[year].keys():
            data_point.append({"y": data_by_year[year][key], "legendText": key})
        dataDict["dataPoints"] = data_point
        data[year] = json.dumps(dataDict)
    return data

if __name__ == '__main__':
    app.debug = True
    app.run()