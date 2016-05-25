from openpyxl import load_workbook
import json


def input_file():
    f = load_workbook(filename='static\Debt Affordability Study Data.xlsx',
                      data_only=True)
    ws = f["Bonds Outstanding (Grandular)"]
    col_title = {}
    for first_row in ws['A1':'M1']:
        for cell in first_row:
            col_title[cell.column] = cell.value

    info_by_row = {}
    for num, row in enumerate(ws.iter_rows(row_offset=1)):
        row_dict = {}
        for cell in row:
            if cell.value is not None:
                row_dict[col_title[cell.column]] = cell.value
        if len(row_dict) != 0:
            info_by_row[num] = row_dict
    return info_by_row


def chart_info_by_year(info):
    info_by_year = default_sort_by_year(info)
    data = {}

    for year in info_by_year.keys():
        dataDict = {"type": "pie",
                    "showInLegend": True,
                    "indexLabel": "{y}"}
        data_point = []
        for pledge in info_by_year[year].keys():
            data_point.append({"y": info_by_year[year][pledge], "legendText": pledge})
        dataDict["dataPoints"] = data_point
        data[year] = json.dumps(dataDict)
    return data


def default_sort_by_year(info):
    data = {}
    for item in info.values():
        if item['DatedDate'].year not in data:
            data[item['DatedDate'].year] = {item['RevenuePledge']: item['Series Amount']}
        else:
            if item['RevenuePledge'] not in data[item['DatedDate'].year].keys():
                data[item['DatedDate'].year][item['RevenuePledge']] = item['Series Amount']
            else:
                data[item['DatedDate'].year][item['RevenuePledge']] += item['Series Amount']
    return data
